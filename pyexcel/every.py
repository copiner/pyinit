'''
author wdaonngg
every
'''

import openpyxl

import math

from util import s_filter


def filter_excel(book_name, sheet_name):

    workbook = openpyxl.load_workbook(book_name)
    sheet = workbook[sheet_name]

    lclass = ['01','02','03','04','05','06','07','08','09','10','11','12']
    #lclass = ['01']
    head1 = []#first line
    head2 = []# second line

    for row in sheet.rows:
        trow = [] #清空

        if row[0].value == '信息':#第一行表头
            for cell in row:
                head1.append(cell.value)

        elif row[0].value == '学号': #第二行表头
            for cell in row:
                head2.append(cell.value)

            #创建sheet 添加表头
            for sname in lclass:
                workbook.create_sheet(sname)
                workbook[sname].append(head1)
                workbook[sname].append(head2)
        else:
            #sheet 添加内容
            for cell in row:
                trow.append(cell.value)

            c = row[3].value # int print(isinstance(c,int))

            if c:
                if c > 9:
                    workbook[str(c)].append(trow)
                else:
                    workbook['0'+str(c)].append(trow)


    workbook.save(filename=book_name)

def cal_excel(book_name,sheet_name):

    workbook = openpyxl.load_workbook(book_name)

    #sheet.max_row-2
    sheet = workbook[sheet_name]

    t_stu = sheet.max_row-2 #总人数 total stu
    t_stu70 = math.ceil((sheet.max_row-2)*0.7) #前70%总人数

    #总分
    t_sco = 0 #总分 score
    t_aver = 0 # 平均分 total average

    a_stu = 0 #A人数
    a_ratio = 0 #A率
    b_stu = 0 #B人数
    d_stu = 0 #D人数
    e_stu = 0 #E人数

    ab_stu = 0 #AB人数
    ab_ratio = 0 #AB率

    de_stu = 0 #DE人数
    de_ratio = 0 #DE率

    t_sco70 = 0 #人数前70%总分
    t_aver70 = 0 #总人数前70%平均分

    a_stu70 = 0 #前70% A人数
    b_stu70 = 0 #前70% B人数
    ab_stu70 = 0 #前70%  AB人数
    ab_ratio70 = 0 #前70%  AB率

    #语文 chinese
    ct_sco = 0 #语文总分 score
    ct_aver = 0 # 平均分 total average

    ca_stu = 0 #A人数
    cb_stu = 0 #B人数
    cd_stu = 0 #D人数
    ce_stu = 0 #E人数

    ca_ratio = 0 #A率

    cab_stu = 0 #AB人数
    cab_ratio = 0 #AB率

    cde_stu = 0 #DE人数
    cde_ratio = 0 #DE率

    cab_stu70 = 0 #AB人数
    cab_ratio70 = 0 #AB率

    ct_sco70 = 0 #人数前70%总分
    ct_aver70 = 0 #人数前70%平均分

    ca_stu70 = 0 #前70% A人数
    cb_stu70 = 0 #前70% B人数
    cab_stu70 = 0 #前70%  AB人数
    cab_ratio70 = 0 #前70%  AB率

    #数学 math
    mt_sco = 0 #数学总分 score
    mt_aver = 0 # 平均分 total average

    ma_stu = 0 #A人数
    mb_stu = 0 #B人数
    md_stu = 0 #D人数
    me_stu = 0 #E人数

    ma_ratio = 0 #A率

    mab_stu = 0 #AB人数
    mab_ratio = 0 #AB率

    mde_stu = 0 #DE人数
    mde_ratio = 0 #DE率

    mab_stu70 = 0 #AB人数
    mab_ratio70 = 0 #AB率

    mt_sco70 = 0 #人数前70%总分
    mt_aver70 = 0 #人数前70%平均分

    ma_stu70 = 0 #前70% A人数
    mb_stu70 = 0 #前70% B人数
    mab_stu70 = 0 #前70%  AB人数
    mab_ratio70 = 0 #前70%  AB率

    #英语 english
    et_sco = 0 #英语总分 score
    et_aver = 0 # 平均分 total average

    ea_stu = 0 #A人数
    eb_stu = 0 #B人数
    ed_stu = 0 #D人数
    ee_stu = 0 #E人数

    ea_ratio = 0 #A率

    eab_stu = 0 #AB人数
    eab_ratio = 0 #AB率

    ede_stu = 0 #DE人数
    ede_ratio = 0 #DE率

    eab_stu70 = 0 #AB人数
    eab_ratio70 = 0 #AB率

    et_sco70 = 0 #人数前70%总分
    et_aver70 = 0 #人数前70%平均分

    ea_stu70 = 0 #前70% A人数
    eb_stu70 = 0 #前70% B人数
    eab_stu70 = 0 #前70%  AB人数
    eab_ratio70 = 0 #前70%  AB率

    #科学science
    st_sco = 0 #科学总分 score
    st_aver = 0 # 平均分 total average

    sa_stu = 0 #A人数
    sb_stu = 0 #B人数
    sd_stu = 0 #D人数
    se_stu = 0 #E人数

    sa_ratio = 0 #A率

    sab_stu = 0 #AB人数
    sab_ratio = 0 #AB率

    sde_stu = 0 #DE人数
    sde_ratio = 0 #DE率

    sab_stu70 = 0 #AB人数
    sab_ratio70 = 0 #AB率

    st_sco70 = 0 #人数前70%总分
    st_aver70 = 0 #人数前70%平均分

    sa_stu70 = 0 #前70% A人数
    sb_stu70 = 0 #前70% B人数
    sab_stu70 = 0 #前70%  AB人数
    sab_ratio70 = 0 #前70%  AB率

    #体育健康 health
    ht_sco = 0 #语文总分 score
    ht_aver = 0 # 平均分 total average

    ha_stu = 0 #A人数
    hb_stu = 0 #B人数
    hd_stu = 0 #D人数
    he_stu = 0 #E人数

    ha_ratio = 0 #A率

    hab_stu = 0 #AB人数
    hab_ratio = 0 #AB率

    hde_stu = 0 #DE人数
    hde_ratio = 0 #DE率

    hab_stu70 = 0 #AB人数
    hab_ratio70 = 0 #AB率

    ht_sco70 = 0 #人数前70%总分
    ht_aver70 = 0 #人数前70%平均分

    ha_stu70 = 0 #前70% A人数
    hb_stu70 = 0 #前70% B人数
    hab_stu70 = 0 #前70%  AB人数
    hab_ratio70 = 0 #前70%  AB率


    #print(isinstance('-', str))
    for row in sheet.iter_rows(min_row=3):

        t_sco = round(t_sco, 1) + s_filter(row[4].value)
        ct_sco = round(ct_sco, 1) + s_filter(row[8].value)
        mt_sco = round(mt_sco, 1) + s_filter(row[12].value)
        et_sco = round(et_sco, 1) + s_filter(row[16].value)
        st_sco = round(st_sco, 1) + s_filter(row[20].value)
        ht_sco = round(ht_sco, 1) + s_filter(row[24].value)

        #总分
        if row[5].value == 'A':
            a_stu += 1
        if row[5].value == 'B':
            b_stu += 1
        if row[5].value == 'D':
            d_stu += 1
        if row[5].value == 'E':
            e_stu += 1
        #语文
        if row[9].value == 'A':
            ca_stu += 1
        if row[9].value == 'B':
            cb_stu += 1
        if row[9].value == 'D':
            cd_stu += 1
        if row[9].value == 'E':
            ce_stu += 1
        #数学
        if row[13].value == 'A':
            ma_stu += 1
        if row[13].value == 'B':
            mb_stu += 1
        if row[13].value == 'D':
            md_stu += 1
        if row[13].value == 'E':
            me_stu += 1
        #英语
        if row[17].value == 'A':
            ea_stu += 1
        if row[17].value == 'B':
            eb_stu += 1
        if row[17].value == 'D':
            ed_stu += 1
        if row[17].value == 'E':
            ee_stu += 1
        #科学
        if row[21].value == 'A':
            sa_stu += 1
        if row[21].value == 'B':
            sb_stu += 1
        if row[21].value == 'D':
            sd_stu += 1
        if row[21].value == 'E':
            se_stu += 1
        #体育健康
        if row[25].value == 'A':
            ha_stu += 1
        if row[25].value == 'B':
            hb_stu += 1
        if row[25].value == 'D':
            hd_stu += 1
        if row[25].value == 'E':
            he_stu += 1

    for row in sheet.iter_rows(min_row=3, max_row=t_stu70+2):
        t_sco70 = round(t_sco, 1) + s_filter(row[4].value)#总分
        ct_sco70 = round(ct_sco70, 1) + s_filter(row[8].value)#语文
        mt_sco70 = round(mt_sco70, 1) + s_filter(row[12].value)
        et_sco70 = round(et_sco70, 1) + s_filter(row[16].value)
        st_sco70 = round(st_sco70, 1) + s_filter(row[20].value)
        ht_sco70 = round(ht_sco70, 1) + s_filter(row[24].value)
        #总分
        if row[5].value == 'A':
            a_stu70 += 1
        if row[5].value == 'B':
            b_stu70 += 1
        #语文
        if row[9].value == 'A':
            ca_stu70 += 1
        if row[9].value == 'B':
            cb_stu70 += 1
        #数学
        if row[13].value == 'A':
            ma_stu70 += 1
        if row[13].value == 'B':
            mb_stu70 += 1
        #英语
        if row[17].value == 'A':
            ea_stu70 += 1
        if row[17].value == 'B':
            eb_stu70 += 1
        #科学
        if row[21].value == 'A':
            sa_stu70 += 1
        if row[21].value == 'B':
            sb_stu70 += 1
        #体育健康
        if row[25].value == 'A':
            ha_stu70 += 1
        if row[25].value == 'B':
            hb_stu70 += 1



    # 精度问题TODO
    t_aver = round(t_sco/t_stu, 1)
    t_aver70 = round(t_sco70/t_stu70, 1)
    a_ratio = round(a_stu/t_stu, 2)

    ab_stu = a_stu+b_stu
    ab_ratio = round(ab_stu/t_stu, 2)

    de_stu = d_stu + e_stu
    de_ratio = round(de_stu/t_stu, 2)

    ab_stu70 = a_stu70 + b_stu70
    ab_ratio70 = round(ab_stu70/t_stu, 2)

    t_list = [
        "",t_stu,"","",
        t_aver,"",t_stu70,t_aver70,"",
        a_stu,a_ratio,"",ab_stu,ab_ratio,"",de_stu,de_ratio,"",
        ab_stu70,ab_ratio70,""
        ]


    # 语文
    ct_aver = round(ct_sco/t_stu, 1)
    ct_aver70 = round(ct_sco70/t_stu70, 1)
    ca_ratio = round(ca_stu/t_stu, 2)

    cab_stu = ca_stu+cb_stu
    cab_ratio = round(cab_stu/t_stu, 2)

    cde_stu = cd_stu + ce_stu
    cde_ratio = round(cde_stu/t_stu, 2)

    cab_stu70 = ca_stu70 + cb_stu70
    cab_ratio70 = round(cab_stu70/t_stu, 2)

    c_list = [
        "",t_stu,"","",
        ct_aver,"",t_stu70,ct_aver70,"",
        ca_stu,ca_ratio,"",cab_stu,cab_ratio,"",cde_stu,cde_ratio,"",
        cab_stu70,cab_ratio70,""
        ]
    # 数学
    mt_aver = round(mt_sco/t_stu, 1)
    mt_aver70 = round(mt_sco70/t_stu70, 1)
    ma_ratio = round(ma_stu/t_stu, 2)

    mab_stu = ma_stu+mb_stu
    mab_ratio = round(mab_stu/t_stu, 2)

    mde_stu = md_stu + me_stu
    mde_ratio = round(mde_stu/t_stu, 2)

    mab_stu70 = ma_stu70 + mb_stu70
    mab_ratio70 = round(mab_stu70/t_stu, 2)

    m_list = [
        "",t_stu,"","",
        mt_aver,"",t_stu70,mt_aver70,"",
        ma_stu,ma_ratio,"",mab_stu,mab_ratio,"",mde_stu,mde_ratio,"",
        mab_stu70,mab_ratio70,""
        ]
    # 英语
    et_aver = round(et_sco/t_stu, 1)
    et_aver70 = round(et_sco70/t_stu70, 1)
    ea_ratio = round(ea_stu/t_stu, 2)

    eab_stu = ea_stu+eb_stu
    eab_ratio = round(eab_stu/t_stu, 2)

    ede_stu = ed_stu + ee_stu
    ede_ratio = round(ede_stu/t_stu, 2)

    eab_stu70 = ea_stu70 + eb_stu70
    eab_ratio70 = round(eab_stu70/t_stu, 2)

    e_list = [
        "",t_stu,"","",
        et_aver,"",t_stu70,et_aver70,"",
        ea_stu,ea_ratio,"",eab_stu,eab_ratio,"",ede_stu,ede_ratio,"",
        eab_stu70,eab_ratio70,""
        ]


    # 科学
    st_aver = round(st_sco/t_stu, 1)
    st_aver70 = round(st_sco70/t_stu70, 1)
    sa_ratio = round(sa_stu/t_stu, 2)

    sab_stu = sa_stu+sb_stu
    sab_ratio = round(sab_stu/t_stu, 2)

    sde_stu = sd_stu + se_stu
    sde_ratio = round(sde_stu/t_stu, 2)

    sab_stu70 = sa_stu70 + sb_stu70
    sab_ratio70 = round(sab_stu70/t_stu, 2)

    s_list = [
        "",t_stu,"","",
        st_aver,"",t_stu70,st_aver70,"",
        sa_stu,sa_ratio,"",sab_stu,sab_ratio,"",sde_stu,sde_ratio,"",
        sab_stu70,sab_ratio70,""
        ]

    # 体育健康
    ht_aver = round(ht_sco/t_stu, 1)
    ht_aver70 = round(ht_sco70/t_stu70, 1)
    ha_ratio = round(ha_stu/t_stu, 2)

    hab_stu = ha_stu+hb_stu
    hab_ratio = round(hab_stu/t_stu, 2)

    hde_stu = hd_stu + he_stu
    hde_ratio = round(hde_stu/t_stu, 2)

    hab_stu70 = ha_stu70 + hb_stu70
    hab_ratio70 = round(hab_stu70/t_stu, 2)

    h_list = [
        "",t_stu,"","",
        ht_aver,"",t_stu70,ht_aver70,"",
        ha_stu,ha_ratio,"",hab_stu,hab_ratio,"",hde_stu,hde_ratio,"",
        hab_stu70,hab_ratio70,""
        ]

    # print(t_list)
    # print(c_list)
    # print(m_list)
    # print(e_list)
    # print(s_list)
    # print(h_list)

    return {
        "t":t_list,
        "c":c_list,
        "m":m_list,
        "e":e_list,
        "s":s_list,
        "h":h_list
    }


def write_excel(book_name):

    workbook = openpyxl.load_workbook(book_name)


    lclass = ['01','02','03','04','05','06','07','08','09','10','11','12']

    dclass = {
        "01":dict(),
        "02":dict(),
        "03":dict(),
        "04":dict(),
        "05":dict(),
        "06":dict(),
        "07":dict(),
        "08":dict(),
        "09":dict(),
        "10":dict(),
        "11":dict(),
        "12":dict()
    }
    #创建sheet 添加表头
    for sname in lclass:
        dclass[sname] = cal_excel(book_name,sname)


    #计算
    lhead = [
        "班级","学籍人数","参考人数","",
        "平均分","平均分名次","班级前70%人数","班级前70%平均分","班级前70%平均分名次",
        "A人数","A率","A率名次","AB人数","AB率","AB率名次","DE人数","DE率","DE率名次",
        "班级前70%AB人数","班级前70%AB率","班级前70%AB率名次"]

    workbook.create_sheet("汇总")

    lcollect = ["班主任-t","语文-c","数学-m","英语-e","科学-s","体育与健康-h"]
    for litem in lcollect:
        lk = litem.split("-")
        lhead[3] = lk[0]
        workbook["汇总"].append(lhead)
        for sname in lclass:
            workbook["汇总"].append(dclass[sname][lk[1]])
        workbook["汇总"].append([])

    #保存
    workbook.save(filename=book_name)


book_name = "every.xlsx";
sheet_name = '成绩';

#filter_excel(book_name, sheet_name)
write_excel(book_name)
